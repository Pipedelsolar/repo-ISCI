import pandas as pd
from datetime import datetime, timedelta
from difflib import SequenceMatcher
import matplotlib.pyplot as plt
import os
from pathlib import Path
import seaborn as sns
import numpy as np
from lifelines.utils import to_long_format
from lifelines.utils import add_covariate_to_timeline
from lifelines import CoxTimeVaryingFitter
from statsmodels.distributions.empirical_distribution import ECDF
from scipy.stats import ks_2samp
from scipy import stats
from matplotlib.ticker import MultipleLocator
import matplotlib.ticker as mtick
from lifelines import KaplanMeierFitter
import warnings
from preproces_prod3 import *
from openpyxl.utils import get_column_letter
from scipy.stats import norm
import gspread
from oauth2client.service_account import ServiceAccountCredentials
warnings.filterwarnings("ignore")

path_actual = Path.cwd()
path_data = path_actual.parent / 'Data'
path_nirsecl = path_actual.parent.parent/'Nirse_cl' / 'Data'

with open(path_data/'lista_ruts_cardio.pkl', 'rb') as f:
    lista_ruts_cardio = pickle.load(f)

filtro_cardio = 'RUN.isin(@lista_ruts_cardio)'
filtro_prematuro = 'RUN.isin(@lista_ruts_preterms)'
filtro_any = '(RUN.isin(@lista_ruts_cardio)) | (RUN.isin(@lista_ruts_preterms))'

df_pf = pd.read_csv(path_data / "df_pf.csv")
df_historic = pd.read_csv(path_data / "df_historic.csv")
df_historic_cardio = pd.read_csv(path_data / "df_historic_cardio.csv")
df_historic_prematuro = pd.read_csv(path_data / "df_historic_prematuro.csv")
pf_nac = pd.read_csv(path_data / "pf_nac.csv")
pf_all = (pd.read_csv(path_data / "pf_all.csv").assign(
        FECHA_NAC = lambda x: pd.to_datetime(x.FECHA_NAC, format = 'mixed'),
        year_nac = lambda x: x.FECHA_NAC.dt.year,
        month_nac = lambda x: x.FECHA_NAC.dt.month,
        elegibilidad_alt = lambda df: (df.year_nac + (df.month_nac >= 10).astype(int)),
        season = lambda x: np.where(x.month_nac.between(4, 9, inclusive='both'), 'in_season','pre_season'),
        ))
nac_filtered_vrs = pd.read_csv(path_data / "nac_filtered_vrs.csv")
nacimientos_per_year_cardio = pd.read_csv(path_data / "nacimientos_per_year_cardio.csv")
nacimientos_per_year_prematuro = pd.read_csv(path_data / "nacimientos_per_year_prematuro.csv")
nacimientos_per_year_any = pd.read_csv(path_data / "nacimientos_per_year_any.csv")

dict_nac_filt = {'cardio': nacimientos_per_year_cardio, 
                 'prematuro': nacimientos_per_year_prematuro, 
                 'any': nacimientos_per_year_any}

tables = []
tables_gr = []


def tabla_impacto_final(filtro):
    df_historic = (pd.read_csv(path_data / "df_historic.csv")
                   .assign(fechaIng = lambda x: pd.to_datetime(x.fechaIng, format='%Y-%m-%d'), 
                           FECHA_NAC = lambda x: pd.to_datetime(x.FECHA_NAC, format='%Y-%m-%d'),
                           #VIVO = 'SI',
                           #FECHA_DEF = pd.NA,
                            ))
    
    criterio_pali_1 = (
    pf_all
    .assign(
        criterio_pali_normal = lambda x: ((x.RUN.isin(lista_ruts_cardio)) | (x.SEMANAS<32) | (x.PESO<1500)).astype(int),
        pali = lambda x: np.where(x.criterio_pali_normal==1, 1, 
                                  np.where((x.year_nac==2023) & (x.month_nac.between(7, 9, inclusive='both')) & (x.SEMANAS<=34) & ((x.PESO<2500)), 1, 0))
            )
    .query(filtro)
    .query('(24<=SEMANAS<=42) & (500<PESO<=p_99999_lognormal)')
    .query('SEXO!=9')
    .RUN
    )

    df_nac_per_year_filt_pali = (
        count_nacs_filter(pf_all, nac_filtered_vrs, '(RUN.isin(@pali_ruts))',pali_ruts=criterio_pali_1)[['elegibilidad_alt','nacs_filtred']]
        .rename(columns={'nacs_filtred':'nacimientos','elegibilidad_alt':'year'})
    )

    egresos_2024 = (
        df_historic
        [~df_historic.RUN.isin(df_historic.query('fechaIng < FECHA_NAC').RUN.unique())]
        .assign(fechaIng = lambda x: pd.to_datetime(x.fechaIng, format='%Y-%m-%d'),
                FECHA_NAC = lambda x: pd.to_datetime(x.FECHA_NAC, format='%Y-%m-%d'),
                ola_enfermedad=lambda df: df.fechaIng.dt.month.between(4, 9),
                elegibilidad_inyear=lambda df: np.where((df.elegibilidad == df.year) & (df.ola_enfermedad),'first_season',
                                                    np.where(((df.elegibilidad + 2 == df.year) & (df.ola_enfermedad)) | 
                                                            ((df.elegibilidad + 1 == df.year) & (df.ola_enfermedad) & (df.season == "pre_season")) ,'second_season', #& (df.season == "in_season")
                                                            'discart')),
        # VIVO_mix = lambda df: np.where((df['VIVO']=='SI') | (df['VIVO'].isna()) & (df['FECHA_DEF'].isna()) , 'SI', 'NO')
        )
        [~df_historic.RUN.isin(df_historic[df_historic['fechaIng'] <= df_historic.FECHA_NAC + pd.DateOffset(days=7)].RUN.unique())]  
        .query('year==2024')
        .query('2022<=elegibilidad<=2024')
        .query('RUN.isin(@criterio_pali_1)')
        .sort_values(by='fechaIng')
        .drop_duplicates(subset=['RUN'], keep='first')
       #.merge(pf_all[['RUN','season']],how='left',on='RUN')
        .query('ola_enfermedad')
        #.query('(VIVO_mix=="SI") | ((VIVO_mix=="NO") & (ingreso_mayor_1_mes>=7))')
        .groupby(['elegibilidad_inyear','year'])
        .VRS1.size()
        .unstack(0)
        .reset_index()
    )

    print(egresos_2024)

    egresos_second_season_2024 = egresos_2024.second_season.values[0]

    print('casos vrs "second_season" considerada:',egresos_second_season_2024)
    
    vrs_eleg_2024 = egresos_2024.first_season.values[0] #df_f_vrs.query('fechaIng_vrs<="2024-09-30"').query(filtro).event_vrs.sum() 

    VRS_table = round(
        df_historic
        [~df_historic.RUN.isin(df_historic.query('fechaIng < FECHA_NAC').RUN.unique())]
        .query('RUN.isin(@criterio_pali_1)')
    # .merge(pf_all[['RUN','season']],how='left',on='RUN')
        .assign(fechaIng = lambda x: pd.to_datetime(x.fechaIng, format='%Y-%m-%d'),
                ola_enfermedad=lambda df: df.fechaIng.dt.month.between(4, 9),
                elegibilidad_inyear=lambda df: np.where((df.elegibilidad == df.year) & (df.ola_enfermedad),'first_season',
                                                    np.where(((df.elegibilidad + 2 == df.year) & (df.ola_enfermedad)) | ((df.elegibilidad + 1 == df.year) & (df.ola_enfermedad) & (df.season == "pre_season")) ,'second_season', #& (df.season == "in_season")
                                                            'discart'))   
        )
        .groupby(['elegibilidad_inyear','year'])
        .VRS1.size()
        .unstack(0)
        .reset_index()
        .drop(columns='discart')
        .merge(df_nac_per_year_filt_pali, on='year', how='left') ###############################
        .merge(df_historic.query('RUN.isin(@criterio_pali_1)').groupby('year', as_index=False).size().rename(columns={'size':'egresos_total'}),on='year',how='left')
        .assign(first_season_per = lambda x: x.first_season * 100000/ x.nacimientos,
                second_season_per = lambda x: x.second_season * 100000/ x.nacimientos,
                ratio = lambda x: x.first_season_per / x.second_season_per )
        .drop(columns=['egresos_total','first_season','second_season'])
        .query('year==2022 | year==2023')
        .set_index('year')
        .T
    ,2)

    nac_24 = df_nac_per_year_filt_pali.query('year==2024').nacimientos.values[0]
    print('los nacimientos 2024 del grupo:',filtro,'son...',nac_24)
    
    tabale_3_proxy = round(
        VRS_table.T.reset_index()
        .assign(vrs_noeleg_2024_per = lambda x: egresos_second_season_2024* 100000/ nac_24, #x.nacimientos, 
                exp_rate_per = lambda x: x.vrs_noeleg_2024_per * x.ratio, 
                exp_numer_cases = lambda x: (egresos_second_season_2024 * x.ratio) * x.nacimientos / nac_24, # AQUÍ ANTES NO NORMALIZABA POR NACS #x.exp_rate_per * nacidos_2024 / 100000,
                averted_num_cases = lambda x: x.exp_numer_cases - (vrs_eleg_2024 * x.nacimientos / nac_24), # AQUÍ ANTES NO NORMALIZABA POR NACS
                relative_reduction = lambda x: 100*x.averted_num_cases/x.exp_numer_cases, # NO INFLUYE LA NORMALIZACIÓN POR NACIMIENTOS (SE CANCELAN)
                averted_num_cases_per = lambda x: x.averted_num_cases * 1000 /x.nacimientos, #/ nac_24,#
                number_needed_to_immunise = lambda x: (1000/x.averted_num_cases_per).astype(int) #np.where(x.averted_num_cases_per>0, (1000/x.averted_num_cases_per).astype(int), 'b')
        )
        .drop(columns=['vrs_noeleg_2024_per'])
    , 2)
    
    main_analisis = pd.DataFrame([tabale_3_proxy.drop(columns=['first_season_per','second_season_per','ratio']).mean()])

    df_tomain = tabale_3_proxy.drop(columns=['first_season_per','second_season_per','ratio'])

    q1 = df_tomain.quantile(0.25).round(2)
    q3 = df_tomain.quantile(0.75).round(2)
    sd = df_tomain.std().round(2)
    med = df_tomain.mean().round(2)

    # formatted = med.astype(str) + " (" + round(q1,2).astype(str) + "-" + round(q3,2).astype(str) + ")"

    formatted = med.astype(str) + " (" + round(sd,2).astype(str) + ")"

    main_analisis = pd.DataFrame([formatted])
    to_loc = main_analisis.year.values[0]
    main_analisis.loc[main_analisis.year==to_loc, 'year'] = 'Main Analysis (sd)'

    Tabla_3_impacto = (
        pd.concat([tabale_3_proxy, main_analisis], ignore_index=True)
        .set_index('year')
        .T
        .reset_index()
        .replace([np.inf, -np.inf], np.nan)
        .fillna('')
        .query('index!="nacimientos"')
    )


    Names_col = ['First RSV season', 'third+ RSV season', 'Ratio (first:third+)', 'Expected rate per 100000',
                'Expected number of cases','Averted number of cases','Relative reduction (%)',
                'Averted number of cases per 1000','Number needed to immunise']

    Tabla_3_impacto = Tabla_3_impacto.assign(index = Names_col)[['index',2022,2023, 'Main Analysis (sd)']]

    na_row = pd.DataFrame([[np.nan] * len(Tabla_3_impacto.columns)], columns=Tabla_3_impacto.columns)

    # Divide el DataFrame original en dos partes e inserta la fila
    Tabla_3_impacto = pd.concat([Tabla_3_impacto.iloc[:3], na_row, Tabla_3_impacto.iloc[3:]], ignore_index=True)
    Tabla_3_impacto.loc[3, 'index'] = 'Estimates for the 2024 RSV season'
    Tabla_3_impacto = pd.concat([Tabla_3_impacto.iloc[:0], na_row, Tabla_3_impacto.iloc[0:]], ignore_index=True)
    Tabla_3_impacto.loc[0, 'index'] = 'RSV hospitalisation rate per 100000'

    Tabla_3_impacto = Tabla_3_impacto.replace([np.inf, -np.inf], np.nan).fillna('')
    
    return Tabla_3_impacto

#################################################################################################################### SECOND SIN PALI + THIRD ########################################################################


Tabla_3_impacto = tabla_impacto_final('pali==1')

with pd.ExcelWriter(path_data/"Tablas_Impacto_Final_2.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    Tabla_3_impacto.to_excel(writer, sheet_name="Pali_impact_final_rev")

tables.append(("Palivizumab-Criteria",Tabla_3_impacto))
tables_gr.append(("Palivizumab-Criteria",Tabla_3_impacto))
########################################################################## CARDIO vs lo mismo #########################################################

Tabla_3_impacto = tabla_impacto_final('(RUN.isin(@lista_ruts_cardio))')

with pd.ExcelWriter(path_data/"Tablas_Impacto_Final_2.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    Tabla_3_impacto.to_excel(writer, sheet_name="Cardio_impact_final_rev")

tables.append(("Congenital-Heart-Disease",Tabla_3_impacto))

########################################################################## SP vs lo mismo #########################################################

Tabla_3_impacto = tabla_impacto_final('(SEMANAS<32) | (PESO<1500)')

with pd.ExcelWriter(path_data/"Tablas_Impacto_Final_2.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    Tabla_3_impacto.to_excel(writer, sheet_name="Sp_impact_final_rev")
    
tables.append(("Extreme Prematurity",Tabla_3_impacto))

########################################################################## Preterms - Pali vs lo mismo #########################################################

Tabla_3_impacto = tabla_impacto_final('(pali==0) & (SEMANAS<=35)') #  '((32<=SEMANAS<=34) & (PESO>2500)) | ((PESO>=1500) & (SEMANAS==35)) & ~(RUN.isin(@lista_ruts_cardio))'

with pd.ExcelWriter(path_data/"Tablas_Impacto_Final_2.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    Tabla_3_impacto.to_excel(writer, sheet_name="preterm_sin_pali_impact_final_rev")
    
#tables.append(("Prematurity out of Palivizumab-Criteria",Tabla_3_impacto))    
tables_gr.append(("Prematurity out of Palivizumab-Criteria",Tabla_3_impacto))

prepared = []
for name, tbl in tables:
    if 'index' in tbl.columns:
        tbl2 = tbl.set_index('index')
    else:
        tbl2 = tbl.copy()
    prepared.append((name, tbl2))

# 2) Concatenar horizontalmente como MultiIndex en columnas
#    Cada bloque de columnas recibirá el nombre de su origen

df_combined = pd.concat(
    {name: tbl for name, tbl in prepared},
    axis=1
)

# 3) Guardar en un único sheet de Excel, usando el índice 'index'
output_path = path_data / "Tablas_Impacto_Final_2.xlsx"
with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    df_combined.to_excel(
        writer,
        sheet_name="Subgroups_Pali",
        index_label='index'
    )
        # Ajustar ancho de cada columna al contenido
    worksheet = writer.sheets["Subgroups_Pali"]
    for i, col_cells in enumerate(worksheet.columns, 1):
        max_length = 0
        for cell in list(col_cells)[1:]:  # Ignorar la primera fila
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                continue
        col_letter = get_column_letter(i)
        worksheet.column_dimensions[col_letter].width = max_length + 1   
        
Tabla_3_impacto = tabla_impacto_final('(pali==1) | (SEMANAS<=35)')

with pd.ExcelWriter(path_data/"Tablas_Impacto_Final_2.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    Tabla_3_impacto.to_excel(writer, sheet_name="High_risk_impact_final_rev")

#tables.append(("High-Risk",Tabla_3_impacto))
tables_gr.append(("High-Risk",Tabla_3_impacto))
tables_gr = [tables_gr[-1]] + tables_gr[:-1]
prepared = []
for name, tbl in tables_gr:
    if 'index' in tbl.columns:
        tbl2 = tbl.set_index('index')
    else:
        tbl2 = tbl.copy()
    prepared.append((name, tbl2))

df_combined = pd.concat(
    {name: tbl for name, tbl in prepared},
    axis=1
)

# 3) Guardar en un único sheet de Excel, usando el índice 'index'
output_path = path_data / "Tablas_Impacto_Final_2.xlsx"
with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    df_combined.to_excel(
        writer,
        sheet_name="Impacto_High_Risk_rev",
        index_label='index'
    )
        # Ajustar ancho de cada columna al contenido
    worksheet = writer.sheets["Impacto_High_Risk"]
    for i, col_cells in enumerate(worksheet.columns, 1):
        max_length = 0
        for cell in list(col_cells)[1:]:  # Ignorar la primera fila
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                continue
        col_letter = get_column_letter(i)
        worksheet.column_dimensions[col_letter].width = max_length + 1   

